import io
import tkinter.messagebox
import tkinter.filedialog
import base64
import os
from ico1 import img
from tkinter import *
from openpyxl import Workbook


######以下部分为华为端口分析程序函数#######


def hwpa(filepath):
    with open(filepath, 'rt',
              encoding='utf-8') as default_file:
        all_config = re.sub(r'#\s*\n', r'#\n', default_file.read())

    ports_list = list(all_config.split('\n#\n'))
    pcount = 1
    portresult = []

    while pcount < len(ports_list) - 1:
        port_config = ports_list[pcount]
        port_name = port_config.split()[1]
        if ("undo shutdown" in port_config) or (re.match(r'[a-zA-Z]+\d+/*\d*/*\d*/*\.\d+',
                                                      port_name) is not None) or (("Eth-Trunk" in port_name) and ("shutdown" not in port_config)):
            # 下行汇聚接口带有相关域配置的话，摘出相应的端口名，描述和域名
            if "bas" in port_config and "#" in port_config:
                domain = ((port_config.split('#')[1]).split('\n')[1]).split()[-1]
                if domain == "layer2-subscriber":
                    domain = "空"
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "description" in port_config.split('\n')[line_current]:
                        descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                    line_current = line_current + 1

                port_tp = ''
                port_tu = ''
                lines = io.StringIO(port_config)
                for line in lines.readlines():
                    line = line.strip()
                    if "traffic-policy" in line and "trust upstream" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                        port_tu = line.split()[-1]
                    elif "traffic-policy" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                    elif "trust upstream" in line:
                        port_tu = line.split()[-1]

                domain_port = [port_name, descr, port_tp, port_tu, "", "所在域为:%s" % domain]
                portresult.append(domain_port)

            # 下行端口配置了组播业务的话，摘取端口号，描述并记录业务类型为组播业务
            elif "igmp" in port_config or "pim" in port_config:
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "description" in port_config.split('\n')[line_current]:
                        descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                    line_current = line_current + 1

                port_tp = ''
                port_tu = ''
                lines = io.StringIO(port_config)
                for line in lines.readlines():
                    line = line.strip()
                    if "traffic-policy" in line and "trust upstream" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                        port_tu = line.split()[-1]
                    elif "traffic-policy" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                    elif "trust upstream" in line:
                        port_tu = line.split()[-1]

                multicast_port = [port_name, descr,  port_tp, port_tu, "组播业务", ""]
                portresult.append(multicast_port)

            # 下行口配置vpn的话，摘取端口号，描述和配置的VPN
            elif "vpn-instance" in port_config:
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "vpn-instance" in port_config.split('\n')[line_current]:
                        vpn_name = port_config.split('\n')[line_current].split()[-1]
                        line_count = len(port_config.split('\n'))
                        line_current = 0
                        while line_current < line_count:
                            if "description" in port_config.split('\n')[line_current]:
                                descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                            line_current = line_current + 1

                        port_tp = ''
                        port_tu = ''
                        lines = io.StringIO(port_config)
                        for line in lines.readlines():
                            line = line.strip()
                            if "traffic-policy" in line and "trust upstream" in line:
                                port_tp = line.split()[-2] + ' ' + line.split()[-1]
                                port_tu = line.split()[-1]
                            elif "traffic-policy" in line:
                                port_tp = line.split()[-2] + ' ' + line.split()[-1]
                            elif "trust upstream" in line:
                                port_tu = line.split()[-1]

                        vpn_port = [port_name, descr, port_tp, port_tu, "", "调用的vpn为:%s" % vpn_name]
                        portresult.append(vpn_port)
                    line_current = line_current + 1

            # 上下行其他端口，只截取端口号和描述
            else:
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "description" in port_config.split('\n')[line_current]:
                        descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                    line_current = line_current + 1

                port_tp = ''
                port_tu = ''
                lines = io.StringIO(port_config)
                for line in lines.readlines():
                    line = line.strip()
                    if "traffic-policy" in line and "trust upstream" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                        port_tu = line.split()[-1]
                    elif "traffic-policy" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                    elif "trust upstream" in line:
                        port_tu = line.split()[-1]
                other_eth = [port_name, descr, port_tp, port_tu, "", ""]
                portresult.append(other_eth)

        pcount = pcount + 1

    # 生成输出excel
    hwwb = Workbook()

    hwportsheet = hwwb.active

    hwportsheet.title = "华为端口使用现状"

    # 编写sheet的列头：

    hwportsheet['A1'].value = '端口名'
    hwportsheet['B1'].value = '端口描述'
    hwportsheet['C1'].value = '已有traffic-policy'
    hwportsheet['D1'].value = '已有标记信任'
    hwportsheet['E1'].value = '业务类型'
    hwportsheet['F1'].value = '备注'

    # 以下为批量转化list型变量内容到xlsx文件对应位置

    for row1 in range(2, len(portresult) + 2):  # 写入数据
        for col1 in range(1, len(portresult[row1 - 2]) + 1):
            _ = hwportsheet.cell(row=row1, column=col1, value=str(portresult[row1 - 2][col1 - 1]))
    devname = re.split(r'[\\/]',filepath)[-1].split()[0]
    hwwb.save(filename="%s 端口分析结果.xlsx" % devname)


######以下部分为阿朗7750_service端口分析程序函数#######


def acsa(filepath):
    with open(filepath, 'rt', encoding='utf-8') as default_file:
        all_config = default_file.read()

    v_or_i_config_mid = list(
        re.split(r'(vprn\s+\d+\s+customer\s+\d+\s+create|ies\s+\d+\s+customer\s+\d+\s+create)', all_config))
    v_or_i_count = 1
    portresult = []
    while v_or_i_count < len(v_or_i_config_mid):
        v_or_i_config_mid2 = v_or_i_config_mid[v_or_i_count]
        if re.match(r'vprn\s+\d+\s+customer\s+\d+\s+create|ies\s+\d+\s+customer\s+\d+\s+create',
                    v_or_i_config_mid2) is not None:
            v_or_i_config = v_or_i_config_mid2 + v_or_i_config_mid[v_or_i_count + 1]
            if 'no shutdown' in v_or_i_config:
                v_or_i_name = ' '.join(v_or_i_config.split('\n')[0].split()[0:4])
                v_or_i_descr = ''
                if "description" in v_or_i_config.split('\n')[1]:
                    v_or_i_descr = v_or_i_config.split('\n')[1].split(maxsplit=1)[1]
                interface_config_mid = list(re.split(r'(interface\s+\S*\s+create)', v_or_i_config))
                interface_count = 1
                while interface_count < len(interface_config_mid):
                    interface_mid2 = interface_config_mid[interface_count]
                    if re.match(r'interface\s+\S*\s+create', interface_mid2) is not None:
                        interface_config = interface_mid2 + interface_config_mid[interface_count + 1]
                        interface_name = interface_config.split('\n')[0].split()[1]
                        interface_descr = ''
                        if "description" in interface_config.split('\n')[1]:
                            interface_descr = interface_config.split('\n')[1].split(maxsplit=1)[1]
                        sap_config_mid = list(re.split(r'(sap\s+\S*\s+create)', interface_config))
                        sap_count = 1
                        while sap_count < len(sap_config_mid):
                            sap_config_mid2 = sap_config_mid[sap_count]
                            if re.match(r'sap\s+\S*\s+create', sap_config_mid2) is not None:
                                sap_config = sap_config_mid2 + sap_config_mid[sap_count + 1]
                                sap_name = sap_config.split('\n')[0].split()[1]
                                ingress_qos_policy = ''
                                egress_qos_policy = ''
                                if re.search(r'ingress\s*\n\s*qos\s+\d+', sap_config) is not None and re.search(
                                        r'egress\s*\n\s*qos\s+\d+', sap_config) is not None:
                                    ingress_qos_policy = re.split(r'(ingress\s*\n\s*qos\s+\d+)', sap_config)[1].split()[
                                        2]
                                    egress_qos_policy = re.split(r'(egress\s*\n\s*qos\s+\d+)', sap_config)[1].split()[2]

                                elif re.search(r'ingress\s*\n\s*qos\s+\d+', sap_config) is not None:
                                    ingress_qos_policy = re.split(r'(ingress\s*\n\s*qos\s+\d+)', sap_config)[1].split()[
                                        2]
                                elif re.search(r'egress\s*\n\s*qos\s+\d+', sap_config) is not None:
                                    egress_qos_policy = re.split(r'(egress\s*\n\s*qos\s+\d+)', sap_config)[1].split()[2]

                                al_service_result = [v_or_i_name, v_or_i_descr, interface_name, interface_descr,
                                                     sap_name,
                                                     ingress_qos_policy,
                                                     egress_qos_policy]
                                portresult.append(al_service_result)
                                sap_count = sap_count + 2

                            else:
                                sap_count = sap_count + 1

                        interface_count = interface_count + 2

                    else:
                        interface_count = interface_count + 1

                v_or_i_count = v_or_i_count + 2

            else:
                v_or_i_count = v_or_i_count + 1

        else:
            v_or_i_count = v_or_i_count + 1


    # 生成输出excel
    hwwb = Workbook()

    hwportsheet = hwwb.active

    hwportsheet.title = "AC 7750 service接口调研结果"

    # 编写sheet的列头：
    hwportsheet['A1'].value = '服务名'
    hwportsheet['B1'].value = '服务描述'
    hwportsheet['C1'].value = '端口名'
    hwportsheet['D1'].value = '端口描述'
    hwportsheet['E1'].value = 'sap名称'
    hwportsheet['F1'].value = 'sap ingress'
    hwportsheet['G1'].value = 'sap engress'
    hwportsheet['H1'].value = '备注'

    # 以下为批量转化list型变量内容到xlsx文件对应位置

    for row1 in range(2, len(portresult) + 2):  # 写入数据
        for col1 in range(1, len(portresult[row1 - 2]) + 1):
            _ = hwportsheet.cell(row=row1, column=col1, value=str(portresult[row1 - 2][col1 - 1]))

    devname = re.split(r'[\\/]', filepath)[-1].split()[0]
    hwwb.save(filename="%s 端口分析结果.xlsx" % devname)


def run_hwpa():
    hw_path = tkinter.filedialog.askopenfilename()
    if hw_path is not '':
        hwpa(hw_path)
        tkinter.messagebox.showinfo("提示", "华为端口分析完毕！")
    else:
        return


def run_acsa():
    ac_path = tkinter.filedialog.askopenfilename()
    if ac_path is not '':
        acsa(ac_path)
        tkinter.messagebox.showinfo("提示", "阿朗7750 service端口分析完毕！")
    else:
        return


def about_message():
    message_content = ("Author：IC\n"
                       "Version：1.4\n"
                       "使用方法：\n"
                       "华为设备：摘取disp cu interface信息，根据端口描述和协议配置，分析端口业务类型，并输出成excel\n"
                       "阿朗7750：摘取config service\n"
                       "info | match \"ies|vprn|vpls|interface|sap|ingress|egress|qos|description|shutdown\" expression信息，根据配置进行解析并输出成excel")
    tkinter.messagebox.showinfo("关于", message_content,)


class App(object):
    def __init__(self, master):
        self.com1 = Button(master, text='华为设备端口调研', command=run_hwpa, width=25)
        self.com1.pack()
        self.com2 = Button(master, text='7750 service端口调研', command=run_acsa, width=25)
        self.com2.pack()
        self.com3 = Button(master, text='关于', command=about_message, width=25)
        self.com3.pack()


root = Tk()
root.title('端口调研程序')
root.geometry("300x100")
with open("tmp.ico","wb+") as tmp:
    tmp.write(base64.b64decode(img))

root.iconbitmap("tmp.ico")
os.remove("tmp.ico")


app = App(root)
root.mainloop()
