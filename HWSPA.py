import io
import tkinter.messagebox
import tkinter.filedialog
import base64
import os
from ico1 import img
from tkinter import *
from openpyxl import Workbook

def hwpa(filepath):
    with open(filepath, 'rt',
              encoding='utf-8') as default_file:
        all_config = re.sub(r'#\s*\n', r'#\n', default_file.read())

    ports_list = list(all_config.split('\n#\n'))
    pcount = 1
    portresult = []

    while pcount < len(ports_list) - 1:
        port_config = ports_list[pcount]
        port_name = port_config.split()[1]
        if ("undo shutdown" in port_config) or (re.match(r'[a-zA-Z]+\d+/*\d*/*\d*/*\.\d+',
                                                      port_name) is not None) or ("shutdown" not in port_config):
            # 下行汇聚接口带有相关域配置的话，摘出相应的端口名，描述和域名
            if "bas" in port_config and "#" in port_config:
                domain = ((port_config.split('#')[1]).split('\n')[1]).split()[-1]
                if domain == "layer2-subscriber":
                    domain = "空"
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "description" in port_config.split('\n')[line_current]:
                        descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                    line_current = line_current + 1

                port_tp = ''
                port_tu = ''
                lines = io.StringIO(port_config)
                for line in lines.readlines():
                    line = line.strip()
                    if "traffic-policy" in line and "trust" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                        port_tu = line.split()[-2] + ' ' + line.split()[-1]
                    elif "traffic-policy" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                    elif "trust" in line:
                        port_tu = line.split()[-2] + ' ' + line.split()[-1]

                domain_port = [port_name, descr, port_tp, port_tu, "", "所在域为:%s" % domain]
                portresult.append(domain_port)

            # 下行端口配置了组播业务的话，摘取端口号，描述并记录业务类型为组播业务
            elif "igmp" in port_config or "pim" in port_config:
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "description" in port_config.split('\n')[line_current]:
                        descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                    line_current = line_current + 1

                port_tp = ''
                port_tu = ''
                lines = io.StringIO(port_config)
                for line in lines.readlines():
                    line = line.strip()
                    if "traffic-policy" in line and "trust" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                        port_tu = line.split()[-2] + ' ' + line.split()[-1]
                    elif "traffic-policy" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                    elif "trust" in line:
                        port_tu = line.split()[-2] + ' ' + line.split()[-1]

                multicast_port = [port_name, descr,  port_tp, port_tu, "组播业务", ""]
                portresult.append(multicast_port)

            # 下行口配置vpn的话，摘取端口号，描述和配置的VPN
            elif "vpn-instance" in port_config:
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "vpn-instance" in port_config.split('\n')[line_current]:
                        vpn_name = port_config.split('\n')[line_current].split()[-1]
                        line_count = len(port_config.split('\n'))
                        line_current = 0
                        while line_current < line_count:
                            if "description" in port_config.split('\n')[line_current]:
                                descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                            line_current = line_current + 1

                        port_tp = ''
                        port_tu = ''
                        lines = io.StringIO(port_config)
                        lines = io.StringIO(port_config)
                        for line in lines.readlines():
                            line = line.strip()
                            if "traffic-policy" in line and "trust" in line:
                                port_tp = line.split()[-2] + ' ' + line.split()[-1]
                                port_tu = line.split()[-2] + ' ' + line.split()[-1]
                            elif "traffic-policy" in line:
                                port_tp = line.split()[-2] + ' ' + line.split()[-1]
                            elif "trust" in line:
                                port_tu = line.split()[-2] + ' ' + line.split()[-1]

                        vpn_port = [port_name, descr, port_tp, port_tu, "", "调用的vpn为:%s" % vpn_name]
                        portresult.append(vpn_port)
                    line_current = line_current + 1

            # 上下行其他端口，只截取端口号和描述
            else:
                line_count = len(port_config.split('\n'))
                line_current = 0
                descr = ''
                while line_current < line_count:
                    if "description" in port_config.split('\n')[line_current]:
                        descr = port_config.split('\n')[line_current].split(maxsplit=1)[1]
                    line_current = line_current + 1

                port_tp = ''
                port_tu = ''
                lines = io.StringIO(port_config)
                for line in lines.readlines():
                    line = line.strip()
                    if "traffic-policy" in line and "trust" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                        port_tu = line.split()[-2] + ' ' + line.split()[-1]
                    elif "traffic-policy" in line:
                        port_tp = line.split()[-2] + ' ' + line.split()[-1]
                    elif "trust" in line:
                        port_tu = line.split()[-2] + ' ' + line.split()[-1]
                other_eth = [port_name, descr, port_tp, port_tu, "", ""]
                portresult.append(other_eth)

        pcount = pcount + 1

    # 生成输出excel
    hwwb = Workbook()

    hwportsheet = hwwb.active

    hwportsheet.title = "华为端口使用现状"

    sheet2 = hwwb.create_sheet()

    sheet2.title = "现状及整改建议"

    # 编写sheet的列头：

    hwportsheet['A1'].value = '端口名'
    hwportsheet['B1'].value = '端口描述'
    hwportsheet['C1'].value = '已有traffic-policy'
    hwportsheet['D1'].value = '已有标记信任'
    hwportsheet['E1'].value = '业务类型'
    hwportsheet['F1'].value = '备注'

    sheet2['A1'].value = '现状'
    sheet2['B1'].value = '整改建议'

    # 以下为批量转化list型变量内容到xlsx文件对应位置

    for row1 in range(2, len(portresult) + 2):  # 写入数据
        for col1 in range(1, len(portresult[row1 - 2]) + 1):
            _ = hwportsheet.cell(row=row1, column=col1, value=str(portresult[row1 - 2][col1 - 1]))
    devname = re.split(r'[\\/]',filepath)[-1].split()[0]
    hwwb.save(filename="%s 端口分析结果.xlsx" % devname)

def run_hwpa():
    hw_path = tkinter.filedialog.askopenfilename()
    if hw_path is not '':
        hwpa(hw_path)
        tkinter.messagebox.showinfo("提示", "华为端口分析完毕！")
    else:
        return

class App(object):
    def __init__(self, master):
        self.com1 = Button(master, text='华为设备端口调研', command=run_hwpa, width=25)
        self.com1.pack()


root = Tk()
root.title('端口调研程序')
root.geometry("300x50")
with open("tmp.ico","wb+") as tmp:
    tmp.write(base64.b64decode(img))

root.iconbitmap("tmp.ico")
os.remove("tmp.ico")


app = App(root)
root.mainloop()